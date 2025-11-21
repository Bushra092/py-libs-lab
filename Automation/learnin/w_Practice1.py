
# ## Core Practice Questions
from openpyxl import Workbook, load_workbook
# ### 1. Load and Read

# Write code to:

# * Load `Students.xlsx`
wb = load_workbook('Students.xlsx')
ws = wb.active
# * Print the value in C3
print(ws['C3'].value)
# * Print the value in A1
print(ws['A1'].value)

# ---

# ### 2. Modify and Save

# Write code to:

# * Load `Marks.xlsx`
m_wb = load_workbook('Marks.xlsx')
m_ws = m_wb.active
# * Change B2 to `85`
m_ws['B2'] = 85
# * Save the file
m_wb.save('AutomatesFileCreation/Marks.xlsx')


# ---

# ### 3. Access Sheets

# You have a workbook with sheets:
# * Summary
# * Data
# * Report

# Write code to:

# * Load the file
access_wb = load_workbook('ReportsFile.xlsx')

# * Access the sheet named `Report`
access_ws = access_wb.active
print(access_ws['Report'])
reportFile = access_ws['Report']
# * Print the value in D5
reportFile['D5'] = 96

# ---

# ### 4. Create a New Workbook

# Write code to:

# * Create a new workbook
wb = Workbook()
# * Rename the active sheet to `Employees`
ws = wb.active
wb.title = 'Employees'
# * Add this row:
#   `['ID', 'Name', 'Salary']`
ws.append(['ID', 'Name', 'Salary'])
# * Save it as `employees.xlsx`
wb.save('AutomatesFileCreation/employees.xlsx')
# ---

# ### 5. Append Rows

# Write code that:

# * Loads `Products.xlsx`
# * Appends this list
#   `['P005', 'Keyboard', 850]`
# * Saves the workbook

# ---

# ### 6. Read Multiple Cells

# Write code to print values of:

# * A1
# * B1
# * C1
#   from a sheet named `Sales`

# ---

# ### 7. List Sheets

# Write code to:

# * Load a workbook
# * Print all sheet names
# * Access the second sheet in the list

# ---

# ### 8. Create Sheet

# Write code to:

# * Load `Finance.xlsx`
# * Create a sheet named `2025`
# * Print all sheet names

# ---

# ### 9. Change Cell Value

# Write code to:

# * Load any workbook
# * Change A1 to `"Updated"`
# * Print A1 before and after the change

# ---

# ### 10. Workbook vs load_workbook

# Write code that:

# 1. Creates a new workbook using `Workbook()`
# 2. Loads an existing workbook using `load_workbook()`
 