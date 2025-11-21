import openpyxl  # Imports the openpyxl package
from openpyxl import Workbook, load_workbook  # Imports Workbook class and load_workbook function

from openpyxl.utils import get_column_letter


# load_workbook: opens an existing Excel file and returns a workbook object

# wb = load_workbook('AutomatesFileCreation/Grades.xlsx')  # Load the Excel file
# ws = wb.active  # Select the active sheet

# print(ws['A1'].value)  # Read value from cell A1
# print(ws['B1'].value)  # Read value from cell B1

# # change info in sheet 
# ws['A1'].value  = 'Name'  #*ALso fine ->  ws['A1']  = 'Name'  
# print(ws['A1'].value)  # After value change 

# #* but doesnt save in excel to do so
# # wb.save('AutomatesFileCreation/Grades.xlsx')


# #? Access diff Sheets
# print('\n_____________Access diff Sheets_________________\n')
# print(wb.sheetnames)
# print(wb['Sheet1'])

# #? Create a new Sheet
# print('\n_____________After creatin new Sheets_________________\n')

# wb.create_sheet('S_openpyxl')
# print(wb.sheetnames)

#? Create a new workbook from scratch
# wb = Workbook()  # Make a new Excel file
# ws = wb.active  # Get the active sheet
# ws.title = 'Data'  # Rename the sheet

# # Add rows to the sheet
# ws.append(['Tim', 'is', 'nice', 'man', '!'])
# ws.append(['Tim', 'is', 'nice', 'man', '!'])
# ws.append(['Tim', 'is', 'nice', 'man', '!'])
# ws.append(['Tim', 'is', 'nice', 'man', '!'])
# ws.append(['Tim', 'is', 'nice', 'man', '!'])
# ws.append(['Tim', 'is', 'nice', 'man', '!'])
# ws.append(['end'])  # Add a final row

# # Save the new workbook

# wb.create_sheet('S_openpyxl')
# wb.create_sheet('S_openpyxl')
# print(wb.sheetnames)

# wb.save('AutomatesFileCreation/dummy.xlsx')


#? Accessin multiple cells


wb = load_workbook('AutomatesFileCreation/dummy.xlsx')
# wb = load_workbook('AutomatesFileCreation/Grades.xlsx')

ws = wb.active

# for row in range(1,11):
#     for col in range(1,5):
#     #    char = chr(65 + col)
#        char = get_column_letter(col)
#     #    print(ws[char + str(row)].value)
#        ws[char+str(row)] = char + str(row)

#? Merging cell
# ws.merge_cells("A1:D2")
# ws.unmerge_cells("A1:D1")

#? insert Rows
# ws.insert_rows(7)
# ws.delete_rows(7)

#? insert Columns
# ws.insert_cols(2)
ws.delete_cols(2)
wb.save('AutomatesFileCreation/dummy.xlsx')

 